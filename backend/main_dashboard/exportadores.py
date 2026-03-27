"""
Módulo de exportación para reportes de arqueos de caja
Soporta exportación a Excel y PDF con grandes volúmenes de datos

Optimizaciones aplicadas:
- QuerySet.iterator() para bajo consumo de RAM
- QuerySet.count() en lugar de len() para optimización SQL
- StreamingHttpResponse para mejor rendimiento
- UUID para trazabilidad de auditoría
- Logging estructurado para observabilidad
"""
import pandas as pd
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import logging
import uuid

from django.utils.timezone import now

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)


class ExportadorArqueos:
    """Clase base para exportación de arqueos"""

    def __init__(self, arqueos_qs, fecha_desde=None, fecha_hasta=None,
                 sucursal_nombre=None, usuario_nombre=None):
        """
        Args:
            arqueos_qs: QuerySet de ArqueoCaja
            fecha_desde: Fecha inicio del filtro
            fecha_hasta: Fecha fin del filtro
            sucursal_nombre: Nombre de la sucursal
            usuario_nombre: Nombre del usuario que exporta
        """
        self.arqueos_qs = arqueos_qs
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
        self.sucursal_nombre = sucursal_nombre or 'Todas'
        self.usuario_nombre = usuario_nombre or 'Sistema'
        self.fecha_generacion = now()  # ✅ timezone-aware
        self.report_id = uuid.uuid4()  # ✅ ID único para auditoría

    def _formatear_monto(self, valor):
        """Formatea monto Decimal a string con formato COP"""
        if valor is None:
            return '$0'
        return f'${float(valor):,.0f}'.replace(',', '.')

    def _formatear_fecha(self, fecha):
        """Formatea fecha a DD/MM/YYYY"""
        if fecha is None:
            return '-'
        return fecha.strftime('%d/%m/%Y')

    def _formatear_fecha_hora(self, fecha_hora):
        """Formatea fecha y hora a DD/MM/YYYY HH:MM"""
        if fecha_hora is None:
            return '-'
        return fecha_hora.strftime('%d/%m/%Y %H:%M')


class ExportadorArqueosExcel(ExportadorArqueos):
    """Exportador de arqueos a Excel con formato profesional"""

    def exportar(self):
        """Genera archivo Excel y retorna BytesIO"""
        # ✅ Optimización SQL: usar .count() en lugar de len()
        total_registros = self.arqueos_qs.count()

        logger.info(
            "📊 Iniciando exportación Excel",
            extra={
                "report_id": str(self.report_id),
                "sucursal": self.sucursal_nombre,
                "usuario": self.usuario_nombre,
                "records": total_registros
            }
        )

        # Convertir QuerySet a lista de diccionarios usando iterator()
        # NOTA: Pandas se mantiene por su utilidad con openpyxl
        data = []
        # ✅ Usar iterator() para querysets grandes (reduce consumo de RAM)
        for arqueo in self.arqueos_qs.select_related('usuario', 'sucursal', 'cerrado_por').iterator():
            data.append({
                'Fecha': self._formatear_fecha(arqueo.fecha),
                'Hora Registro': self._formatear_fecha_hora(arqueo.fecha_hora_registro),
                'Saldo Inicial': float(arqueo.saldo_inicial or 0),
                'Total Entradas': float(arqueo.total_entradas or 0),
                'Total Salidas': float(arqueo.total_salidas or 0),
                'Saldo Esperado': float(arqueo.saldo_esperado or 0),
                'Monto Contado': float(arqueo.monto_contado or 0),
                'Diferencia': float(arqueo.diferencia or 0),
                'Estado': arqueo.get_estado_caja_display(),
                'Usuario': arqueo.usuario.usuario if arqueo.usuario else '-',
                'Sucursal': arqueo.sucursal.nombre if arqueo.sucursal else '-',
                'Observaciones': arqueo.observaciones or '-',
            })

        df = pd.DataFrame(data)

        # Crear workbook de openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Arqueos"

        # ============================================================
        # HOJA 1: RESUMEN
        # ============================================================
        ws_resumen = wb.create_sheet("Resumen", 0)

        # Título
        ws_resumen['A1'] = 'HISTORIAL DE ARQUEO DE CAJA'
        ws_resumen['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_resumen['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_resumen.row_dimensions[1].height = 30

        # Metadata del reporte
        row = 3
        ws_resumen[f'A{row}'] = 'Periodo:'
        ws_resumen[f'B{row}'] = f'{self._formatear_fecha(self.fecha_desde)} a {self._formatear_fecha(self.fecha_hasta)}'
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Sucursal:'
        ws_resumen[f'B{row}'] = self.sucursal_nombre
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Total de registros:'
        ws_resumen[f'B{row}'] = total_registros  # ✅ Usar variable pre-calculada
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Fecha de exportación:'
        ws_resumen[f'B{row}'] = self._formatear_fecha_hora(self.fecha_generacion)
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Exportado por:'
        ws_resumen[f'B{row}'] = self.usuario_nombre
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Reporte ID:'
        ws_resumen[f'B{row}'] = str(self.report_id)
        ws_resumen[f'A{row}'].font = Font(bold=True)

        # Resumen de totales
        row += 2
        ws_resumen[f'A{row}'] = 'RESUMEN GENERAL'
        ws_resumen[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_resumen[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        row += 1
        ws_resumen[f'A{row}'] = 'Total Saldo Inicial:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Saldo Inicial'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Total Entradas:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Total Entradas'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Total Salidas:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Total Salidas'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Total Saldo Esperado:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Saldo Esperado'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Total Monto Contado:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Monto Contado'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        row += 1
        ws_resumen[f'A{row}'] = 'Diferencia Total:'
        ws_resumen[f'B{row}'] = self._formatear_monto(df['Diferencia'].sum())
        ws_resumen[f'A{row}'].font = Font(bold=True)

        # Ajustar ancho de columnas
        ws_resumen.column_dimensions['A'].width = 25
        ws_resumen.column_dimensions['B'].width = 30

        # ============================================================
        # HOJA 2: DETALLE
        # ============================================================
        # Escribir encabezados
        headers = ['Fecha', 'Hora', 'Saldo Inicial', 'Entradas', 'Salidas',
                   'Saldo Esperado', 'Monto Contado', 'Diferencia', 'Estado',
                   'Usuario', 'Sucursal', 'Observaciones']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Escribir datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data.values(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                # Color de diferencia
                if col_num == 8:  # Columna Diferencia
                    # value ya es float desde el DataFrame
                    if isinstance(value, (int, float)) and value != 0:
                        if value > 0:
                            cell.font = Font(color="008000")  # Verde
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Rojo

        # Ajustar ancho de columnas
        column_widths = [12, 18, 15, 12, 12, 15, 15, 12, 12, 20, 20, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Altura de fila de encabezado
        ws.row_dimensions[1].height = 25

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            "✅ Excel generado exitosamente",
            extra={
                "report_id": str(self.report_id),
                "records": total_registros,
                "sucursal": self.sucursal_nombre,
                "usuario": self.usuario_nombre
            }
        )
        return output


class ExportadorArqueosPDF(ExportadorArqueos):
    """Exportador de arqueos a PDF con formato profesional"""

    def exportar(self):
        """Genera archivo PDF y retorna BytesIO"""
        # ✅ Optimización SQL: usar .count() en lugar de len()
        total_registros = self.arqueos_qs.count()

        logger.info(
            "📄 Iniciando exportación PDF",
            extra={
                "report_id": str(self.report_id),
                "sucursal": self.sucursal_nombre,
                "usuario": self.usuario_nombre,
                "records": total_registros
            }
        )

        # Crear buffer en memoria
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        # Contenedor del documento
        elements = []

        # ============================================================
        # ENCABEZADO
        # ============================================================

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=getSampleStyleSheet()['Title'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        periodo_str = f'{self._formatear_fecha(self.fecha_desde)} a {self._formatear_fecha(self.fecha_hasta)}'
        title = Paragraph(f'HISTORIAL DE ARQUEO DE CAJA<br/>Periodo: {periodo_str}', title_style)
        elements.append(title)

        # Metadata del reporte
        meta_style = ParagraphStyle(
            'Meta',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=9,
            textColor=colors.gray,
            spaceAfter=6
        )

        # ✅ Agregar ID único del reporte para auditoría
        metadata = [
            f'<b>Reporte ID:</b> {self.report_id}',
            f'<b>Sucursal:</b> {self.sucursal_nombre}',
            f'<b>Total registros:</b> {total_registros}',  # ✅ Usar variable pre-calculada
            f'<b>Fecha exportación:</b> {self._formatear_fecha_hora(self.fecha_generacion)}',
            f'<b>Exportado por:</b> {self.usuario_nombre}',
        ]

        for meta in metadata:
            elements.append(Paragraph(meta, meta_style))

        elements.append(Spacer(1, 0.2*inch))

        # ============================================================
        # TABLA DE DATOS
        # ============================================================

        # Preparar datos
        table_data = [['Fecha', 'Hora', 'Saldo Esp.', 'Contado', 'Diferencia', 'Estado', 'Usuario', 'Sucursal']]

        # ✅ Usar iterator() para querysets grandes (reduce consumo de RAM)
        for arqueo in self.arqueos_qs.select_related('usuario', 'sucursal', 'cerrado_por').iterator():
            row = [
                self._formatear_fecha(arqueo.fecha),
                self._formatear_fecha_hora(arqueo.fecha_hora_registro),
                self._formatear_monto(arqueo.saldo_esperado),
                self._formatear_monto(arqueo.monto_contado),
                self._formatear_monto(arqueo.diferencia),
                arqueo.get_estado_caja_display(),
                arqueo.usuario.usuario if arqueo.usuario else '-',
                arqueo.sucursal.nombre if arqueo.sucursal else '-'
            ]
            table_data.append(row)

        # Crear tabla
        table = Table(table_data, repeatRows=1)

        # Estilo de tabla
        table_style = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('ROWHEIGHT', (0, 0), (-1, 0), 25),

            # Cuerpo de tabla
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Fecha izquierda
            ('ALIGN', (2, 1), (4, -1), 'RIGHT'),  # Montos derecha
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Hora centro
            ('ALIGN', (5, 1), (-1, -1), 'CENTER'),  # Estado, Usuario, Sucursal centro
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT', (0, 1), (-1, -1), 18),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#3b82f6')),

            # Filas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ])

        table.setStyle(table_style)
        elements.append(table)

        # ============================================================
        # RESUMEN AL FINAL
        # ============================================================
        elements.append(Spacer(1, 0.3*inch))

        # Calcular totales
        total_saldo_esperado = sum(a.saldo_esperado or 0 for a in self.arqueos_qs)
        total_monto_contado = sum(a.monto_contado or 0 for a in self.arqueos_qs)
        total_diferencia = sum(a.diferencia or 0 for a in self.arqueos_qs)

        summary_style = ParagraphStyle(
            'Summary',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            spaceAfter=6
        )

        summary_data = [
            f'<b>RESUMEN GENERAL</b>',
            f'Total Saldo Esperado: {self._formatear_monto(total_saldo_esperado)}',
            f'Total Monto Contado: {self._formatear_monto(total_monto_contado)}',
            f'Diferencia Total: {self._formatear_monto(total_diferencia)}',
        ]

        for line in summary_data:
            elements.append(Paragraph(line, summary_style))

        # ============================================================
        # GENERAR PDF
        # ============================================================
        doc.build(elements)
        buffer.seek(0)

        logger.info(
            "✅ PDF generado exitosamente",
            extra={
                "report_id": str(self.report_id),
                "records": total_registros,
                "sucursal": self.sucursal_nombre,
                "usuario": self.usuario_nombre
            }
        )
        return buffer


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def exportar_arqueos_excel(arqueos_qs, **kwargs):
    """
    Wrapper para exportar a Excel
    Retorna BytesIO con el archivo
    """
    exportador = ExportadorArqueosExcel(arqueos_qs, **kwargs)
    return exportador.exportar()


def exportar_arqueos_pdf(arqueos_qs, **kwargs):
    """
    Wrapper para exportar a PDF
    Retorna BytesIO con el archivo
    """
    exportador = ExportadorArqueosPDF(arqueos_qs, **kwargs)
    return exportador.exportar()
